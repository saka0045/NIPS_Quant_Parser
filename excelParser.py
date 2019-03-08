#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Jun 21 14:39:11 2018

@author: m006703
"""

from openpyxl import load_workbook
import os
import argparse


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        '-i', '--inputDir', dest='basedir', required=True,
        help="Path to input directory of NIPS RTFs"
    )
    parser.add_argument(
        '-o', '--outpath', dest='outpath', required=True,
        help="Path to output file"
    )

    args = parser.parse_args()

    basedir = os.path.abspath(args.basedir)
    outpath = os.path.abspath(args.outpath)

    # Add / at the end if it is not included in the paths
    if basedir.endswith("/"):
        basedir = basedir
    else:
        basedir = basedir + "/"

    if outpath.endswith("/"):
        outpath = outpath
    else:
        outpath = outpath + "/"

    # Add all of the RTFs in a list
    file_list = os.listdir(basedir)

    average_size_list, final_quant_list, initial_quant_list, run_name_list, sample_name_list = parse_nips_rtf(basedir,
                                                                                                              file_list)

    result_file = open(outpath + "NIPSQuantResult.txt", 'w')
    result_file.write("RunName\tSampleName\tInitialQuant\tFinalQuant\tAverageSize\n")

    for index in range (0, len(sample_name_list)):
        result_file.write(run_name_list[index] + "\t" + sample_name_list[index] + "\t" + initial_quant_list[index] +
                          "\t" + final_quant_list[index] + "\t" + str(average_size_list[index]) + "\n")

    result_file.close()

    print("Script is done running")
    print("Results are saved at " + outpath + "NipsQuantResult.txt")


def parse_nips_rtf(basedir, file_list):
    """
    Parses NIPS RTFs in basedir to obtain the Initial Quant, Final Quant and Average Size
    Use the NipsQuantBoxPlot.R R script to make a boxplot from this data
    :param basedir:
    :param file_list:
    :return initial and final quant, average size, sample name, NIPS run name per NIPS RTF:
    """
    sample_name_list = []
    initial_quant_list = []
    final_quant_list = []
    average_size_list = []
    run_name_list = []
    for file in file_list:
        # Make sure it's a NIPS RTF
        if file[0:4] == 'NIPS':
            file_path = basedir + file
            # This script will work but the result file will be messed up if it's not in the form of NIPS###_RTF.xlsm
            run_name = file.split('_')[0]
            print("Processing file: " + file_path)
            # Make sure the read_only=True is there, otherwise it takes forever!
            wb = load_workbook(file_path, data_only=True, read_only=True)

            quant_sheet = wb['DNA Quantitation']

            for x in range(4, 999):
                try:
                    # Sometimes the NTC is NPP-NTC or NTC...
                    if 'NTC' in quant_sheet.cell(row=x, column=2).value:
                        break
                    else:
                        sample_name = quant_sheet.cell(row=x, column=2).value
                        sample_name_list.append(sample_name)
                        try:
                            initial_quant = ("%.2f" % quant_sheet.cell(row=x, column=5).value)
                        except TypeError:
                            initial_quant = 'NA'
                        initial_quant_list.append(initial_quant)
                        try:
                            final_quant = ("%.2f" % quant_sheet.cell(row=x, column=7).value)
                        except TypeError:
                            final_quant = 'NA'
                        final_quant_list.append(final_quant)
                        # Gather the average size for every 8 samples
                        if ((x - 4) % 8) == 0:
                            try:
                                average_size = quant_sheet.cell(row=x, column=9).value
                            except TypeError:
                                average_size = 'NA'
                        average_size_list.append(average_size)
                        run_name_list.append(run_name)
                # Skip if there are any blank lines between samples and NTC
                except TypeError:
                    continue
    return average_size_list, final_quant_list, initial_quant_list, run_name_list, sample_name_list


if __name__ == "__main__":
    main()
